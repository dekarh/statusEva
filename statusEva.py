# -*- coding: utf-8 -*-
# Заполняем поле статуса в монго конвертируя его из Excel

import sys, argparse
from _datetime import datetime, timedelta, date
import time
import os
from mysql.connector import MySQLConnection, Error
from collections import OrderedDict
import openpyxl
from pymongo import MongoClient
import psycopg2

from lib import read_config, s

STATUSES = {
'BANK REFUSAL': 430,
'APPROVED': 140,
'CLIENT REFUSAL': 440,
'ISSUED': 210,
'CANCEL': 450,
'EXPIRED': 460
}


st = """
DECISION (решение) - тут будет несколько статусов 
Cancel - закрыта
Decline - 
Expired - истек срок действия решения Банка
Issued - выдан
Pending - ожидание

нас интересует статус из этого столбца 
Cancel, Issued,Expired

Bank refusal - Отказ банка
Approved - Одобрен
Client refusal - Отказ клиента
issued - Выдан

NONE	0	Новая заявка
QUEUED	100	Заявка отправлена в очередь
CONFIRM	110	Введен СМС код
RETRY	120	Запрошена повторная СМС
PROCESSING	130	В процессе
APPROVED	140	Одобрена
PRE_APPROVED	150	Предварительно одобрена
COMPLETED	160	Заявка заполнена
DONE	200	Завершено
ISSUED	210	Займ выдан
DOUBLE_ISSUED	220	Займ выдан повторно
ISSUED_CALLCENTER	230	Займ выдан через call-центр
DELETED	400	Удалена
UNKNOWN	410	Неизвестный статус
TRANSACTION_ERROR	420	Ошибка выгрузки
DENIED	430	Отказ
DEBUG	500	Отладка
DRAFT	510	Отложена
"""
def filter_x00(inp):
    inp = s(inp)
    inp = inp.replace('_x0020_',' ')
    inp = inp.replace('_X0020_',' ')
    while inp.upper().find('_X0') > -1:
        if inp.find('_x0') > -1:
            inp = inp.split('_x0')[0] + inp.split('_x0')[1].split('_')[1]
        else:
            inp = inp.split('_X0')[0] + inp.split('_X0')[1].split('_')[1]
    return inp

if __name__ == '__main__':
    # подключаемся к серверу
    cfg = read_config(filename='anketa.ini', section='Mongo')
    conn = MongoClient('mongodb://' + cfg['user'] + ':' + cfg['password'] + '@' + cfg['ip'] + ':' + cfg['port'] + '/'
                       + cfg['db'])
    # выбираем базу данных
    db = conn.saturn_v
    # выбираем коллекцию документов
    colls = db.Products

    # Sort file names with path
    path = "./"
    file_list = os.listdir(path)
    full_list = [os.path.join(path, i) for i in file_list if i.startswith('Raiffeisen_Finfort_') and i.endswith('.xlsx')]
    xlsxs = sorted(full_list, key = os.path.getmtime)

    for xlsx in xlsxs:
        print('\n', xlsx,'\n')
        wb = openpyxl.load_workbook(filename=xlsx)
        ws = wb[wb.sheetnames[0]]
        wbo = openpyxl.Workbook(write_only=True)
        wso_ish = wbo.create_sheet('Исходный')
        wso_task = wbo.create_sheet('Задание')
        wso_skip_id = wbo.create_sheet('Нет id')
        wso_skip_status = wbo.create_sheet('Нет статуса')
        wso_double = wbo.create_sheet('Два id в одной строке')
        wso_rez = wbo.create_sheet('Результат')
        ids = []
        column_utm_source = -1
        column_approval = -1
        column_remote_id = -1
        column_result = -1
        column_decision = -1
        for i, row in enumerate(ws.values):
            # заполняем вкладку задания
            fields_task = []
            for cell in row:
                fields_task.append(cell)
            wso_task.append(fields_task)
            # определяем колонку в которой id
            if not i:
                wso_skip_status.append(row)
                wso_skip_id.append(row)
                wso_double.append(row)
                for j, cell in enumerate(row):
                    if str(cell).upper() == 'UTM_TERM':
                        column_utm_source = j
                    if str(cell).upper() == 'APPROVAL':
                        column_approval = j
                    if str(cell).upper() == 'REMOTE_ID':
                        column_remote_id = j
                    if cell == 'RESULT':
                        column_result = j
                    if cell == 'DECISION':
                        column_decision = j
            else:
                # Если нет нужной информации - выходим
                if (column_utm_source < 0 and column_remote_id < 0) or (column_approval < 0 and column_result < 0
                                                                        and column_decision < 0):
                    print('Нет колонки с id или колонки со статусом')
                    sys.exit()
                # Если не смогли расшифровать статус - пропускаем строчку
                status = -1
                if column_decision > -1:
                    status = STATUSES.get(filter_x00(row[column_decision]).upper().strip(), -1)
                if status < 0 and column_approval > -1:
                    status = STATUSES.get(filter_x00(row[column_approval]).upper().strip(), -1)
                if status < 0 and column_result > -1:
                    status = STATUSES.get(filter_x00(row[column_result]).upper().strip(), -1)
                if status < 0: # Нет статуса
                    wso_skip_status.append(row)
                    continue
                remote_id = ''
                remote_id_utm = ''
                remote_id_remote = ''
                if column_utm_source > -1 and str(type(row[column_utm_source])).find('str') > -1:
                    agent2remote_id = row[column_utm_source]
                    if len(filter_x00(agent2remote_id)[filter_x00(agent2remote_id).find('_') + 1:].strip()) == 36:
                        remote_id_utm = filter_x00(agent2remote_id)[filter_x00(agent2remote_id).find('_') + 1:].strip()
                        if not colls.find({'remote_id': remote_id_utm}).count():
                            remote_id_utm = ''
                if column_remote_id > -1 and str(type(row[column_remote_id])).find('str') > -1:
                    if len(filter_x00(row[column_remote_id].strip())) == 36:
                        remote_id_remote = row[column_remote_id].strip()
                        if not colls.find({'remote_id': remote_id_remote}).count():
                            remote_id_remote = ''
                if remote_id_remote == '' and remote_id_utm == '': # Нет id
                    wso_skip_id.append(row)
                    continue
                elif remote_id_remote and remote_id_utm: # Два id в одной строке
                    wso_double.append(row)
                    continue
                elif remote_id_utm:
                    remote_id = remote_id_utm
                elif remote_id_remote:
                    remote_id = remote_id_remote
                # заполняем вкладку исходника
                for j, coll in enumerate(colls.find({'remote_id': remote_id})):
                    if not j:
                        fields_ish = []
                        for field in coll.keys():
                            if str(type(coll.get(field))).find('str') < 0 and str(type(coll.get(field))).find(
                                    'int') < 0:
                                fields_ish.append(str(coll.get(field)))
                            else:
                                fields_ish.append(coll.get(field))
                        wso_ish.append(fields_ish)
                # обновляем
                #colls.update({'remote_id': remote_id}, {'$set': {'state_code': status}})
                # заполняем вкладку результата
                for j, coll in enumerate(colls.find({'remote_id': remote_id})):
                    if not j:
                        fields_rez = []
                        for field in coll.keys():
                            if str(type(coll.get(field))).find('str') < 0 and str(type(coll.get(field))).find(
                                    'int') < 0:
                                fields_rez.append(str(coll.get(field)))
                            else:
                                fields_rez.append(coll.get(field))
                        wso_rez.append(fields_rez)
        wbo.save(xlsx.split('Raiffeisen_Finfort_')[0] + 'loaded/' +
                 time.strftime('%Y-%m-%d_%H-%M', time.gmtime(os.path.getmtime(xlsx))) + '_' +
                 xlsx.split('Raiffeisen_Finfort_')[1])
        os.remove(xlsx)


